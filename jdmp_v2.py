import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Border, Side, Alignment

st.title("Judaica Digital Metadata Parser (Prototype)")
#st.header("Importing, Cleaning, Validation, Template Population, Exporting")

# --- upload files ---
urns_file = st.file_uploader("**Upload URNs Excel**", type=["xlsx"])
desc_file = st.file_uploader("**Upload Descriptive Metadata Excel**", type=["xlsx"])

with st.expander("**📤 Optional Uploads (click to expand)**"):
    template_file = st.file_uploader("Upload SharedShelf Template Excel (optional - if none uploaded, will use default SharedShelf template)", type=["xlsx"])

# --- template file handling ---
    # cached function: loads template (or any Excel) file once, then reuses result
    @st.cache_data
    def load_template(path: str):
        return pd.read_excel(path)

    if template_file: # if user uploads a new template
        try:
            template_df = pd.read_excel(template_file)
            st.success(f"Custom Template loaded: {template_df.shape[1]} columns detected")
        except Exception as e:
            st.error(f"**Could not read the uploaded Template: {e}**")
            template_df = None
    else: # fallback to default stored template
        try:
            template_df = load_template("SharedShelf Template.xlsx")
            #st.info("No template uploaded. Using default SharedShelf template.")
            st.success(f"Default SharedShelf Template: {template_df.shape[1]} columns detected")
        except Exception as e:
            st.error(f"**Default Template not found or unreadable: {e}**")
            template_df = None

# --- crediting file handling ---
    crediting_file = st.file_uploader("Upload Crediting-Notes Translation Table (optional)", type=["xlsx"])

    @st.cache_data
    def load_crediting_table(path):
        # read first two columns only and normalize
        df = pd.read_excel(path)
        df = df.iloc[:, :2].copy()
        df.columns = ["source", "notes"]
        df["source"] = df["source"].astype(str).str.strip()
        df["notes"] = df["notes"].astype(str).fillna("").str.strip()
        df = df.dropna(subset=["source"])
        return df

    if crediting_file:  # if user uploads a new table
        try:
            crediting_df = load_crediting_table(crediting_file)
            st.success(f"Custom Crediting-Notes Translation Table loaded: {len(crediting_df)} sources")
        except Exception as e:
            st.error(f"**Could not read the uploaded file: {e}**")
            crediting_df = None
    else:  # fallback to default table
        try:
            crediting_df = load_crediting_table("Notes-Crediting - Translation Table - Column DB.xlsx")
            st.success(f"Default Crediting-Notes Translation Table: {len(crediting_df)} sources")
        except Exception as e:
            st.error(f"**Default file not found or unreadable: {e}**")
            crediting_df = None

# --- URNs file handling ---
missing_selections = []

if urns_file:
    urns_df = pd.read_excel(urns_file)
    st.subheader("URNs")

    # drop rows with NaN or blank FILE-URN
    if "FILE-URN" in urns_df.columns:
        urns_df = urns_df.dropna(subset=["FILE-URN"]).copy()
        urns_df = urns_df[(urns_df["FILE-URN"].astype(str).str.strip() != "")].reset_index(drop=True)
        st.success(f"Cleaned URNs: {len(urns_df)} rows remaining")
    else:
        st.error("**Column 'FILE-URN' not found in URNs file.**")

    urns_cols = urns_df.columns.tolist()
    urns_cols_with_none = [None] + urns_cols

    # select match field
    urns_key_col = st.selectbox("**Select Match Field from URNs Spreadsheet (usually FILE-OSN)**", urns_cols_with_none)

    if urns_key_col is None:
        missing_selections.append("Match Field for URNs Spreadsheet")

# --- URNs image preview utility ---
if urns_file and "FILE-URN" in urns_df.columns:
    with st.expander("🖼️ Preview URN Images (click to expand)"):
        urns_df["FILE-URN"] = urns_df["FILE-URN"].astype(str).str.strip()
        urns_df["image_url"] = "http://nrs.harvard.edu/" + urns_df["FILE-URN"] + "?"
        st.success(f"{len(urns_df)} URNs processed. Preview associated images below.")

        if len(urns_df) == 0:
            st.warning("**No rows to preview after cleaning.**")
            st.stop()

        # keep session index valid and in sync
        if "image_index" not in st.session_state:
            st.session_state.image_index = 0
        st.session_state.image_index = max(0, min(st.session_state.image_index, len(urns_df) - 1))

        idx = st.session_state.image_index
        row = urns_df.iloc[idx]

        st.markdown(f"[Open in browser (if no image below, click to log in)]({row['image_url']})")
        st.markdown(f"**Image {idx + 1} of {len(urns_df)}**")
        st.markdown(f"**URN:** {row['FILE-URN']}")

        # numeric jump box (1-based for users)
        jump_val = st.number_input(
            "Go to image",
            min_value=1,
            max_value=len(urns_df),
            value=idx + 1,
            step=1,
            help="Enter a number to jump directly to that image."
        )
        if (jump_val - 1) != st.session_state.image_index:
            st.session_state.image_index = jump_val - 1
            idx = st.session_state.image_index
            row = urns_df.iloc[idx]

        # side-by-side layout: prev | image | next 
        col_prev, col_img, col_next = st.columns([1, 8, 1])

        with col_prev:
            st.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)
            if st.button("⬅️", width="stretch") and st.session_state.image_index > 0:
                st.session_state.image_index -= 1

        with col_next:
            st.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)
            if st.button("➡️", width="stretch") and st.session_state.image_index < len(urns_df) - 1:
                st.session_state.image_index += 1

        # refresh row after any button click
        st.session_state.image_index = max(0, min(st.session_state.image_index, len(urns_df) - 1))
        idx = st.session_state.image_index
        row = urns_df.iloc[idx]

        with col_img:
            try:
                st.image(row["image_url"], width="stretch")
            except Exception as e:
                st.warning(f"**Could not load image for URN {row['FILE-URN']}: {e}**")

# --- descriptive metadata file handling (relevant selections included) ---
if desc_file:
    desc_df = pd.read_excel(desc_file)
    st.subheader("Descriptive Metadata")

    desc_cols = desc_df.columns.tolist()
    desc_cols_with_none = [None] + desc_cols

    # select types
    metadata_type = st.selectbox("**Select Metadata Type**", [None, "Posters", "Ephemera", "Memorabilia"])
    cataloging_type = st.radio("**Select Cataloging Type**", [ "Full Cataloging", "Provisional Records"], horizontal=True)
    geographic_type = st.selectbox("**Select Geographic Type**", [None, "Israel", "World Judaica"])
    if geographic_type == "World Judaica":
        artstor_country_col = st.selectbox("**Select Country Column from Desc Metadata Spreadsheet**", desc_cols_with_none)
    else:
        artstor_country_col = ""

    # select columns
    desc_key_col = st.selectbox("**Select Match Field from Desc Metadata Spreadsheet**", desc_cols_with_none, index=2)  # default: 2nd column
    desc_title_col = st.selectbox("**Select Title Column from Desc Metadata Spreadsheet**", desc_cols_with_none)
    desc_start_date_col = st.selectbox("**Select Start Date Column from Desc Metadata Spreadsheet**", desc_cols_with_none)
    desc_end_date_col = st.selectbox("**Select End Date Column from Desc Metadata Spreadsheet**", desc_cols_with_none)

    # select general note
    desc_source_type = st.selectbox("**Select Source for General Note / Shareshelf Description**",
                                    [None, "Descriptive Metadata Column", "NO GENERAL NOTE", "OTHER"])
    if desc_source_type == "Descriptive Metadata Column":
        desc_note_col = st.selectbox("Select the Note Column", [None] + desc_cols)
    elif desc_source_type == "OTHER":
        desc_source_text = st.text_area("Enter Custom General Note")

    # check if user made all required selections
    if metadata_type is None:
        missing_selections.append("Metadata Type")
    if geographic_type is None:
        missing_selections.append("Geographic Type")
    if geographic_type == "World Judaica" and artstor_country_col is None:
        missing_selections.append("Artstor Country Column")
    if desc_key_col is None:
        missing_selections.append("Match Field for Descriptive Metadata")
    if desc_title_col is None:
        missing_selections.append("Title Column")
    if desc_start_date_col is None:
        missing_selections.append("Start Date Column")
    if desc_end_date_col is None:
        missing_selections.append("End Date Column")
    if desc_source_type is None:
        missing_selections.append("Source for General Note")
    
    # store choices in session state
    #st.session_state["metadata_type"] = metadata_type
    #st.session_state["geographic_type"] = geographic_type
    #st.session_state["cataloging_type"] = cataloging_type
    #st.session_state["desc_title_col"] = desc_title_col
    #st.session_state["desc_start_date_col"] = desc_start_date_col
    #st.session_state["desc_end_date_col"] = desc_end_date_col

# --- template-related selections ---
if urns_file and desc_file and template_df is not None:
    st.subheader("Template Population")

    # enter creator & subject info
    template_creator = st.text_area("**Enter Creator Information**")
    template_subject = st.text_area("**Enter Subject Information**")

    # select copyright info
    template_rights_type = st.selectbox("**Select Source for Rights**", [None, "STANDARD", "OTHER"])
    if template_rights_type == "STANDARD":
        template_rights_text = "The President and Fellows of Harvard College make no representation that they are the owner of the copyright; any researcher wishing to make use of an image must therefore assume all responsibility for clearing reproduction rights and for any infringement of Title 17 of the United States Code."
    elif template_rights_type == "OTHER":
        template_rights_text = st.text_area("Enter Custom Copyright Information")
    
    # select crediting info
    template_credit_type = None
    template_credit_text = ""

    if crediting_df is not None and not crediting_df.empty:
        crediting_df_source = crediting_df["source"].tolist()

        template_credit_type = st.selectbox(
            "**Select Source for Crediting**",
            [None] + crediting_df_source + ["OTHER"]
        )

        if template_credit_type and template_credit_type != "OTHER":
            credit_df_notes = crediting_df.loc[crediting_df["source"] == template_credit_type, "notes"]
            template_credit_text = next((t for t in credit_df_notes if t and t.strip()), "")
            if not template_credit_text:
                st.warning("**Selected source has no corresponding note in the table.**")
        elif template_credit_type == "OTHER":
            template_credit_text = st.text_area("Enter Custom Crediting Information")
    else:
        st.error("**No valid Crediting-Notes Traslation Table available. Upload one or include the default file in the app repo.**")

    # check if user made all required selections
    if template_rights_type is None:
        missing_selections.append("Source for Rights")
    if template_credit_type is None:
        missing_selections.append("Source for Crediting")

# --- validation ---
if urns_file and desc_file:
    if urns_key_col is not None and desc_key_col is not None:
        urn_keys = set(urns_df[urns_key_col].astype(str).str.strip())
        desc_keys = set(desc_df[desc_key_col].astype(str).str.strip())

        # check row count
        if len(urns_df) != len(desc_df):
            st.warning(f"**Row count mismatch! URNs: {len(urns_df)}, Descriptive Metadata: {len(desc_df)}**")

        # check key sets
        desc_missing_keys = urn_keys - desc_keys
        urns_missing_keys = desc_keys - urn_keys

        if desc_missing_keys or urns_missing_keys:
            st.warning("**Key mismatch detected!**")
            if desc_missing_keys:
                st.write(f"URNs not in Descriptive Metadata: {list(desc_missing_keys)}")
            if urns_missing_keys:
                st.write(f"Descriptive Metadata not in URNs: {list(urns_missing_keys)}")

        # allow override (TBD)
        #override = st.checkbox("Override mismatch warning and proceed")
        #if override:
        #    st.success("Override enabled: You can proceed to populate the template")

    else:
        st.info("Please select Match Fields for validation to run.")

# --- template population pipeline ---
if urns_file and desc_file and template_df is not None:
    st.subheader("Populated SharedShelf Template")

    # initialize blank template aligned to URNs row count
    target_rows = len(urns_df)
    template_out = template_df.head(0).copy()
    template_out = template_out.reindex(range(target_rows)).reset_index(drop=True)

    # category 1: standard fixed values
    try:
        template_out.loc[:, "SSID"] = "NEW"
        template_out.loc[:, "File Count"] = 1
        template_out.loc[:, "Repository[34349]"] = "Judaica Division, Widener Library[9000347138]"
        template_out.loc[:, "Image Repository[34365]"] = "Judaica Division, Widener Library[9000347138]"
        template_out.loc[:, "Send To Harvard[34382]"] = True
        template_out.loc[:, "In House Use Only[34383]"] = False
        template_out.loc[:, "Export Only In Group[34411]"] = False
        
        template_fixed_val_cols = ["SSID", "File Count", "Repository[34349]", "Image Repository[34365]",
                          "Send To Harvard[34382]", "In House Use Only[34383]", "Export Only In Group[34411]"]
    except KeyError as e:
        st.error(f"**Template missing expected column(s) for fixed value population: {e}**")

    # category 2: URNs value population - FILE-URN + FILE-OSN
    try:
        template_out.loc[:, "Filename"] = "drs:" + urns_df["FILE-URN"].astype(str).str.strip()

        urns_df["FILE-OSN_transformed"] = urns_df["OBJ-OSN"].astype(str).str.upper().str.replace("_", "", n=1)
        template_out.loc[:, "Repository Classification Number[34364]"] = urns_df["FILE-OSN_transformed"]
        template_out.loc[:, "Image Classification Number[34369]"] = urns_df["FILE-OSN_transformed"]
        template_out.loc[:, "Repository Number[2560412]"] = urns_df["FILE-OSN_transformed"] + " (classification)"

        template_urns_cols = ["Filename", "Repository Classification Number[34364]", "Image Classification Number[34369]",
                              "Repository Number[2560412]"]
    except KeyError as e:
        st.error(f"**Template missing expected column(s) for URN-related population: {e}**")

    # category 3-1: descriptive metadata population - start/end dates
    if desc_start_date_col is not None and desc_end_date_col is not None:
        start = pd.to_numeric(desc_df[desc_start_date_col], errors="coerce")
        end   = pd.to_numeric(desc_df[desc_end_date_col], errors="coerce")
        template_date_cols = ["Date Description[34341]", "ARTstor Earliest Date[34342]", "ARTstor Latest Date[34343]",
                              "Earliest Date[2560433]", "Latest Date[2560435]"]

        def assign_dates(start, end, template_date_warnings):
            if pd.notna(start) and pd.notna(end) and (start != end):  # both present & different
                if start > end:
                    template_date_warnings.add("**One or more rows have Start Date later than End Date; Start Date used as default.**")
                    return int(start), int(start), int(start), int(start), int(start)
                return f"{int(start)}-{int(end)}", int(start), int(end), int(start), int(end)
            elif pd.notna(start) and pd.notna(end) and (start == end):  # both present & identical
                return int(start), int(start), int(end), int(start), int(end)
            elif pd.notna(start) and pd.isna(end):  # start present, end blank
                return int(start), int(start), int(start), int(start), int(start)
            elif pd.isna(start) and pd.notna(end):  # start blank, end present
                template_date_warnings.add("**One or more rows have blank Start Date; End Date used as default.**")
                return int(end), int(end), int(end), int(end), int(end)
            else:  # both blank
                template_date_warnings.add("**One or more rows have both Start and End Dates missing; defaulted to 1900–2025.**")
                return "1900-2025", 1900, 2025, 1900, 2025
        
        template_date_warnings = set()  # avoid duplicated warnings
        date_values = [assign_dates(s, e, template_date_warnings) for s, e in zip(start, end)]

        for msg in template_date_warnings:
            st.warning(msg)

        date_df = pd.DataFrame(date_values, columns=template_date_cols)
        for col in date_df.columns:
            try:
                template_out[col] = date_df[col]
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Start/End Date Population: {e}**")

    # category 3-2: descriptive metadata population - title
    if (desc_title_col and metadata_type and cataloging_type) is not None:
        if desc_title_col not in desc_df.columns:
            st.error("**Selected Title column not found in Descriptive Metadata.**")
            #st.stop()
        
        titles = desc_df[desc_title_col].astype(str).str.strip()

        # define logic only for posters & ephemera for now
        if metadata_type == "Posters":
            if cataloging_type == "Full Cataloging":
                populated_titles = titles
            elif cataloging_type == "Provisional Records":
                    populated_titles = titles + " - poster (Cataloging in progress)"
            else:
                st.warning("**Unknown Cataloging Type; titles left blank.**")
                populated_titles = ""
        
        if metadata_type == "Ephemera":
            if cataloging_type == "Full Cataloging":
                populated_titles = titles
            elif cataloging_type == "Provisional Records":
                    populated_titles = titles + " - ephemera item (Cataloging in progress)"
            else:
                st.warning("**Unknown Cataloging Type; titles left blank.**")
                populated_titles = ""

        try:
            template_out.loc[:, "Title[34338]"] = populated_titles
        except KeyError as e:
            st.error(f"**Template missing expected column for Title population: {e}**")

    # category 3-3: descriptive metadata population - metadata type-related
    if metadata_type is not None:
        if metadata_type == "Posters":
            try:
                #template_out.loc[:, "Creator[34336]"] = ""
                template_out.loc[:, "Materials/Techniques[34345]"] = "posters"
                template_out.loc[:, "Work Type[34348]"] = "posters"
                template_out.loc[:, "Materials Techniques Note[2560408]"] = "posters"

                template_meta_type_cols = ["Creator[34336]", "Materials/Techniques[34345]", "Work Type[34348]", "Materials Techniques Note[2560408]"]
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Metadata Type-related population: {e}**")
        if metadata_type == "Ephemera":
            try:
                #template_out.loc[:, "Creator[34336]"] = ""
                template_out.loc[:, "Materials/Techniques[34345]"] = "ephemera"
                template_out.loc[:, "Work Type[34348]"] = "ephemera"
                template_out.loc[:, "Materials Techniques Note[2560408]"] = "ephemera"

                template_meta_type_cols = ["Creator[34336]", "Materials/Techniques[34345]", "Work Type[34348]", "Materials Techniques Note[2560408]"]
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Metadata Type-related population: {e}**")

    # category 3-4: descriptive metadata population - general note
    if "Description[34357]" in template_out.columns:
        if desc_source_type is not None:
            if desc_source_type == "Descriptive Metadata Column" and desc_note_col:
                template_out.loc[:, "Description[34357]"] = desc_df[desc_note_col].astype(str).str.strip()
            elif desc_source_type == "NO GENERAL NOTE":
                template_out.loc[:, "Description[34357]"] = ""
            elif desc_source_type == "OTHER" and desc_source_text:
                template_out.loc[:, "Description[34357]"] = desc_source_text
            else:
                st.warning("**Please select a valid General Note source or text.**")
    else:
        st.error("**Template missing expected column for General Note population: 'Description[34357]'**")

    # category 3-5: descriptive metadata population - culture
    if "Culture[34337]" in template_out.columns:
        if geographic_type is not None:
            if geographic_type == "Israel":
                template_out.loc[:, "Culture[34337]"] = "Israeli"
            elif geographic_type == "World Judaica":
                template_out.loc[:, "Culture[34337]"] = "Jewish"
    else:
        st.error("**Template missing expected column for Culture Type population: 'Culture[34337]'**")

    # category 3-6: descriptive metadata population - artstor country
    if "Artstor Country[34356]" in template_out.columns:
        if geographic_type is not None:
            if geographic_type == "Israel":
                template_out.loc[:, "Artstor Country[34356]"] = "Israel"
            elif geographic_type == "World Judaica" and artstor_country_col is not None:
                template_out.loc[:, "Artstor Country[34356]"] = desc_df[artstor_country_col].astype(str).str.strip()
    else:
        st.error("**Template missing expected column for Country Information population: 'Artstor Country[34356]'**")

    # category 4: copyright + crediting info
    if template_rights_type is not None:
        if template_rights_text != "":
            try:
                template_out.loc[:, "Rights[34363]"] = template_rights_text
                template_out.loc[:, "Rights/Access Information[2560402]"] = template_rights_text
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Copyright Information population: {e}**")
        else:
            st.warning("**Please enter Copyright Information.**")
    
    if template_credit_type is not None:
        if template_credit_text != "":
            try:
                template_out.loc[:, "Notes[2560400]"] = template_credit_text
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Crediting Note population: {e}**")
        else:
            st.error("**Crediting Note cannot be blank.**")

    # save intermediate for future categories
    st.session_state["template_out"] = template_out

    # show combined preview of what’s been filled so far
    preview_cols = []
    if "template_fixed_val_cols" in locals():
        preview_cols += template_fixed_val_cols
    if "template_urns_cols" in locals():
        preview_cols += template_urns_cols
    if "populated_titles" in locals():
        preview_cols += ["Title[34338]"]
    if "template_date_cols" in locals():
        preview_cols += template_date_cols
    if "geographic_type" in locals():
        preview_cols += ["Culture[34337]", "Artstor Country[34356]"]
    if "template_meta_type_cols" in locals():
        preview_cols += template_meta_type_cols
    if "desc_source_type" in locals():
        preview_cols += ["Description[34357]"]
    if "template_rights_type" in locals():
        preview_cols += ["Rights[34363]", "Rights/Access Information[2560402]"]
    if "template_credit_type" in locals():
        preview_cols += ["Notes[2560400]"]

    st.dataframe(template_out[preview_cols].head(10), width="stretch")

    # export / download
    if missing_selections:
        st.warning(
            f"**Please select value(s) for {', '.join(missing_selections)} before downloading the populated template.**"
        )

    else:
        # - Excel -
        xlsx_output = io.BytesIO()
        with pd.ExcelWriter(xlsx_output, engine="openpyxl") as writer:
            template_out.to_excel(writer, index=False, sheet_name="Sheet1")

            # styling: borders + top alignment + wrap
            worksheet = writer.sheets["Sheet1"]
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            align_top = Alignment(vertical="top", horizontal="left", wrap_text=True)

            for row in worksheet.iter_rows(
                min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column
            ):
                for cell in row:
                    cell.border = border
                    cell.alignment = align_top

        st.download_button(
            label="Download Populated SharedShelf Template (Excel)",
            data=xlsx_output.getvalue(),
            file_name="JDMP_Populated_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # - CSV -
        # use UTF-8 with BOM so Excel on Windows opens it without mojibake
        csv_bytes = template_out.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label="Download Populated SharedShelf Template (CSV)",
            data=csv_bytes,
            file_name="JDMP_Populated_Template.csv",
            mime="text/csv",
    )

