import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Border, Side

st.title("JDMP Prototype")
st.header("Importing, Cleaning, Validation, Template Population (IP), Exporting")

# --- upload files ---
urns_file = st.file_uploader("Upload URNs Excel", type=["xlsx"])
desc_file = st.file_uploader("Upload Descriptive Metadata Excel", type=["xlsx"])
template_file = st.file_uploader("Upload SharedShelf Template Excel (optional)", type=["xlsx"])

# --- template file handling ---
# cached function: loads template (or any Excel) file once, then reuses result
@st.cache_data
def load_template(path: str):
    return pd.read_excel(path)

if template_file: # if user uploads a new template
    try:
        template_df = pd.read_excel(template_file)
        st.success(f"Custom template loaded: {template_df.shape[1]} columns detected")
    except Exception as e:
        st.error(f"**Could not read the uploaded template: {e}**")
        template_df = None

else: # fallback to default stored template
    try:
        template_df = load_template("SharedShelf Template.xlsx")
        #st.info("No template uploaded. Using default SharedShelf template.")
        st.success(f"Default SharedShelf template: {template_df.shape[1]} columns detected")
    except Exception as e:
        st.error(f"**Default template not found or unreadable: {e}**")
        template_df = None

# --- URNs file handling ---
if urns_file:
    urns_df = pd.read_excel(urns_file)
    st.subheader("URNs")

    # drop rows with NaN or blank FILE-URN
    if "FILE-URN" in urns_df.columns:
        urns_df = urns_df.dropna(subset=["FILE-URN"]).copy()
        urns_df = urns_df[(urns_df["FILE-URN"].astype(str).str.strip() != "")].reset_index(drop=True)
        st.success(f"Cleaned URNs: {len(urns_df)} rows remaining")
    else:
        st.error("**Column 'FILE-URN' not found in URNs file.**")

    urns_cols = urns_df.columns.tolist()

    # select the match field (default = OBJ-OSN)
    urns_default_key_col = "OBJ-OSN"
    if urns_default_key_col in urns_cols:
        urns_default_key_index = urns_cols.index(urns_default_key_col)
    else:
        urns_default_key_index = 0
    urns_key_col = st.selectbox("Select the Match Field", urns_cols, index=urns_default_key_index)

# --- descriptive metadata file handling (relevant selections included) ---
missing_selections = []

if desc_file:
    desc_df = pd.read_excel(desc_file)
    st.subheader("Descriptive Metadata")

    desc_cols = desc_df.columns.tolist()
    desc_cols_with_none = [None] + desc_cols

    # select types
    metadata_type = st.selectbox("Select Metadata Type", [None, "Posters", "Ephemera", "Memorabilia"])
    cataloging_type = st.radio("Select Cataloging Type", [ "Full Cataloging", "Provisional Records"], horizontal=True)
    geographic_type = st.selectbox("Select Geographic Type", [None, "Israel", "World Judaica"])

    # select columns
    desc_key_col = st.selectbox("Select the Match Field", desc_cols, index=1)  # default: 2nd column
    desc_title_col = st.selectbox("Select the Title Column", desc_cols_with_none)
    desc_start_date_col = st.selectbox("Select the Start Date Column", desc_cols_with_none)
    desc_end_date_col = st.selectbox("Select the End Date Column", desc_cols_with_none)

    # select description source
    desc_source_type = st.selectbox("Select the Source for Description",
                                    [None, "Descriptive Metadata Column", "NO DESCRIPTION NOTE", "OTHER"])
    if desc_source_type == "Descriptive Metadata Column":
        desc_note_col = st.selectbox("Select the Note Column", [None] + desc_cols)
    elif desc_source_type == "OTHER":
        desc_source_text = st.text_area("Enter Custom Description Note")

    # check if user made all required selections
    if metadata_type is None:
        missing_selections.append("Metadata Type")
    if geographic_type is None:
        missing_selections.append("Geographic Type")
    if desc_title_col is None:
        missing_selections.append("Title Column")
    if desc_start_date_col is None:
        missing_selections.append("Start Date Column")
    if desc_end_date_col is None:
        missing_selections.append("End Date Column")
    if desc_source_type is None:
        missing_selections.append("Description Source")
    
    # store choices in session state
    #st.session_state["metadata_type"] = metadata_type
    #st.session_state["geographic_type"] = geographic_type
    #st.session_state["cataloging_type"] = cataloging_type
    #st.session_state["desc_title_col"] = desc_title_col
    #st.session_state["desc_start_date_col"] = desc_start_date_col
    #st.session_state["desc_end_date_col"] = desc_end_date_col

# --- template-related selections ---
if template_df is not None:
    # select copyright info
    template_rights_type = st.selectbox("Select Copyright Information", [None, "STANDARD", "OTHER"])
    if template_rights_type == "STANDARD":
        template_rights_text = "The President and Fellows of Harvard College make no representation that they are the owner of the copyright; any researcher wishing to make use of an image must therefore assume all responsibility for clearing reproduction rights and for any infringement of Title 17 of the United States Code."
    elif template_rights_type == "OTHER":
        template_rights_text = st.text_area("Enter Custom Copyright Information")
    
    # select crediting info
    template_credit_type = st.selectbox("Select Crediting Information", [
        None, "231 Lowe", "435 Swibel", "409 Cowett E", "431 Cowett F&J", 
        "436 Cowett W.", "437 Jacobson", "153 Hvd Litt", "OTHER"])
    if template_credit_type == "231 Lowe":
        template_credit_text = "Digitization funded from the income of the Joe and Emily Lowe Foundation Book Fund for Judaica in the Harvard College Library (Fund 560231)."
    elif template_credit_type == "435 Swibel":
        template_credit_text = "Digitization funded from the income of the Howard J. Swibel Library Preservation Fund in the Harvard College Library (Fund 560435)."
    elif template_credit_type == "409 Cowett E":
        template_credit_text = "Digitization funded from the income of the Edward M. Cowett 1951 Memorial Judaica Preservation Fund in the Harvard College Library (Fund 560409)."
    elif template_credit_type == "431 Cowett F&J":
        template_credit_text = "Digitization funded from the income of the Florence and Joseph B. Cowett Memorial Fund for Judaica Preservation in the Harvard College Library (Fund 560431)."
    elif template_credit_type == "436 Cowett W.":
        template_credit_text = "Digitization funded from the income of the Wilbur A. Cowett Judaica Preservation Fund in the Harvard College Library (Fund 560436)."
    elif template_credit_type == "437 Jacobson":
        template_credit_text = "Digitization funded from the income of the Joan Leiman Jacobson Fund for the Preservation of Judaica in the Harvard College Library (Fund 560437)."
    elif template_credit_type == "153 Hvd Litt":
        template_credit_text = "Digitization funded from the income of the Harvard-Littauer Judaica Endowment in the Harvard College Library (Fund 560153)."
    elif template_credit_type == "OTHER":
        template_credit_text = st.text_area("Enter Custom Crediting Information")

    # check if user made all required selections
    if template_rights_type is None:
        missing_selections.append("Copyright Information")
    if template_credit_type is None:
        missing_selections.append("Crediting Information")

# --- validation ---
if urns_file and desc_file:
    if "urns_key_col" in locals() and "desc_key_col" in locals():
        urn_keys = set(urns_df[urns_key_col].astype(str).str.strip())
        desc_keys = set(desc_df[desc_key_col].astype(str).str.strip())

        # check row count
        if len(urns_df) != len(desc_df):
            st.warning(f"**Row count mismatch! URNs: {len(urns_df)}, Descriptive Metadata: {len(desc_df)}**")

        # check key sets
        desc_missing_keys = urn_keys - desc_keys
        urns_missing_keys = desc_keys - urn_keys

        if desc_missing_keys or urns_missing_keys:
            st.warning("**Key mismatch detected!**")
            if desc_missing_keys:
                st.write(f"URNs not in Descriptive Metadata: {list(desc_missing_keys)}")
            if urns_missing_keys:
                st.write(f"Descriptive Metadata not in URNs: {list(urns_missing_keys)}")

        # allow override (TBD)
        #override = st.checkbox("Override mismatch warning and proceed")
        #if override:
        #    st.success("Override enabled: You can proceed to populate the template")

    else:
        st.info("Please select the match fields for validation to run.")

# --- template population pipeline ---
if urns_file and desc_file and template_df is not None:
    st.subheader("Populated SharedShelf Template")

    # initialize blank template aligned to URNs row count
    target_rows = len(urns_df)
    template_out = template_df.head(0).copy()
    template_out = template_out.reindex(range(target_rows)).reset_index(drop=True)

    # category 1: standard fixed values
    try:
        template_out.loc[:, "SSID"] = "NEW"
        template_out.loc[:, "File Count"] = 1
        template_out.loc[:, "Repository[34349]"] = "Judaica Division, Widener Library[9000347138]"
        template_out.loc[:, "Image Repository[34365]"] = "Judaica Division, Widener Library[9000347138]"
        template_out.loc[:, "Send To Harvard[34382]"] = True
        template_out.loc[:, "In House Use Only[34383]"] = False
        template_out.loc[:, "Export Only In Group[34411]"] = False
        
        template_fixed_val_cols = ["SSID", "File Count", "Repository[34349]", "Image Repository[34365]",
                          "Send To Harvard[34382]", "In House Use Only[34383]", "Export Only In Group[34411]"]
    except KeyError as e:
        st.error(f"**Template missing expected column(s) for fixed value population: {e}**")

    # category 2: URNs value population - FILE-URN + FILE-OSN
    try:
        template_out.loc[:, "Filename"] = "drs:" + urns_df["FILE-URN"].astype(str).str.strip()

        urns_df["FILE-OSN_transformed"] = urns_df["OBJ-OSN"].astype(str).str.upper().str.replace("_", "", n=1)
        template_out.loc[:, "Repository Classification Number[34364]"] = urns_df["FILE-OSN_transformed"]
        template_out.loc[:, "Image Classification Number[34369]"] = urns_df["FILE-OSN_transformed"]
        template_out.loc[:, "Repository Number[2560412]"] = urns_df["FILE-OSN_transformed"] + " (classification)"

        template_urns_cols = ["Filename", "Repository Classification Number[34364]", "Image Classification Number[34369]",
                              "Repository Number[2560412]"]
    except KeyError as e:
        st.error(f"**Template missing expected column(s) for URN-related population: {e}**")

    # category 3-1: descriptive metadata population - start/end dates
    if desc_start_date_col is not None and desc_end_date_col is not None:
        start = pd.to_numeric(desc_df[desc_start_date_col], errors="coerce")
        end   = pd.to_numeric(desc_df[desc_end_date_col], errors="coerce")
        template_date_cols = ["Date Description[34341]", "ARTstor Earliest Date[34342]", "Latest Date[34343]",
                              "Earliest Date[2560433]", "Latest Date[2560435]"]

        def assign_dates(start, end, template_date_warnings):
            if pd.notna(start) and pd.notna(end) and (start != end):  # both present & different
                if start > end:
                    template_date_warnings.add("**One or more rows have Start Date later than End Date; Start Date used as default.**")
                    return int(start), int(start), int(start), int(start), int(start)
                return f"{int(start)}-{int(end)}", int(start), int(end), int(start), int(end)
            elif pd.notna(start) and pd.notna(end) and (start == end):  # both present & identical
                return int(start), int(start), int(end), int(start), int(end)
            elif pd.notna(start) and pd.isna(end):  # start present, end blank
                return int(start), int(start), int(start), int(start), int(start)
            elif pd.isna(start) and pd.notna(end):  # start blank, end present
                template_date_warnings.add("**One or more rows have blank Start Date; End Date used as default.**")
                return int(end), int(end), int(end), int(end), int(end)
            else:  # both blank
                template_date_warnings.add("**One or more rows have both Start and End Dates missing; defaulted to 1900–2025.**")
                return "1900-2025", 1900, 2025, 1900, 2025
        
        template_date_warnings = set()  # avoid duplicated warnings
        date_values = [assign_dates(s, e, template_date_warnings) for s, e in zip(start, end)]

        for msg in template_date_warnings:
            st.warning(msg)

        date_df = pd.DataFrame(date_values, columns=template_date_cols)
        for col in date_df.columns:
            try:
                template_out[col] = date_df[col]
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Start/End Date Population: {e}**")

    # category 3-2: descriptive metadata population - title
    if (desc_title_col and metadata_type and cataloging_type and cataloging_type) is not None:
        if desc_title_col not in desc_df.columns:
            st.error("**Selected Title column not found in Descriptive Metadata.**")
            #st.stop()
        
        titles = desc_df[desc_title_col].astype(str).str.strip()

        # define logic only for posters for now
        if metadata_type == "Posters":
            if cataloging_type == "Full Cataloging":
                populated_titles = titles
            elif cataloging_type == "Provisional Records":
                    populated_titles = titles + " - poster (Cataloging in progress)"
            else:
                st.warning("**Unknown Cataloging Type; titles left blank.**")
                populated_titles = ""

        try:
            template_out.loc[:, "Title[34338]"] = populated_titles
        except KeyError as e:
            st.error(f"**Template missing expected column for Title population: {e}**")

    # category 3-3: descriptive metadata population - metadata type-related
    if metadata_type is not None:
        if metadata_type == "Posters":
            try:
                template_out.loc[:, "Creator[34336]"] = ""
                template_out.loc[:, "Materials/Techniques[34345]"] = "posters"
                template_out.loc[:, "Work Type[34348]"] = "posters"
                template_out.loc[:, "Materials Techniques Note[2560408]"] = "posters"

                template_meta_type_cols = ["Creator[34336]", "Materials/Techniques[34345]", "Work Type[34348]", "Materials Techniques Note[2560408]"]
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Metadata Type-related population: {e}**")

    # category 3-4: descriptive metadata population - description
    if "Description[34357]" in template_out.columns:
        if desc_source_type is not None:
            if desc_source_type == "Descriptive Metadata Column" and desc_note_col:
                template_out.loc[:, "Description[34357]"] = desc_df[desc_note_col].astype(str).str.strip()
            elif desc_source_type == "NO DESCRIPTION NOTE":
                template_out.loc[:, "Description[34357]"] = ""
            elif desc_source_type == "OTHER" and desc_source_text:
                template_out.loc[:, "Description[34357]"] = desc_source_text
            else:
                st.warning("**Please select a valid Description source or text.**")
    else:
        st.error("**Template missing expected column for Description population: 'Description[34357]'**")

    # category 4: copyright + crediting info
    if template_rights_type is not None:
        if template_rights_text != "":
            try:
                template_out.loc[:, "Rights[34363]"] = template_rights_text
                template_out.loc[:, "Rights/Access Information[2560402]"] = template_rights_text
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Copyright Information population: {e}**")
        else:
            st.warning("**Please enter Copyright Information.**")
    
    if template_credit_type is not None:
        if template_credit_text != "":
            try:
                template_out.loc[:, "Notes[2560400]"] = template_credit_text
            except KeyError as e:
                st.error(f"**Template missing expected column(s) for Crediting Note population: {e}**")
        else:
            st.error("**Crediting Note cannot be blank.**")

    # save intermediate for future categories
    st.session_state["template_out"] = template_out

    # show combined preview of what’s been filled so far
    preview_cols = []
    if "template_fixed_val_cols" in locals():
        preview_cols += template_fixed_val_cols
    if "template_urns_cols" in locals():
        preview_cols += template_urns_cols
    if "populated_titles" in locals():
        preview_cols += ["Title[34338]"]
    if "template_date_cols" in locals():
        preview_cols += template_date_cols
    if "template_meta_type_cols" in locals():
        preview_cols += template_meta_type_cols
    if "desc_source_type" in locals():
        preview_cols += ["Description[34357]"]
    if "template_rights_type" in locals():
        preview_cols += ["Rights[34363]", "Rights/Access Information[2560402]"]
    if "template_credit_type" in locals():
        preview_cols += ["Notes[2560400]"]

    st.dataframe(template_out[preview_cols].head(10), use_container_width=True)

    # export to Excel
    if missing_selections:
        st.warning(
            f"**Please select value(s) for {', '.join(missing_selections)} before downloading the populated template.**"
        )

    else:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            template_out.to_excel(writer, index=False)

            # apply border to all cells in the exported excel
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column
                                           ):
                for cell in row:
                    cell.border = border
        st.download_button(
            label="Download Populated SharedShelf Template (Excel)",
            data=output.getvalue(),
            file_name="JDMP_Populated_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

